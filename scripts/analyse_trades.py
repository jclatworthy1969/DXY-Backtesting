"""
DXY Backtesting Trade Analysis
Reads Ash Mall and Brice Strebler trade data, filters to last 50 DXY wins,
computes indicator checks, and writes Phase1_TradeAnalysis.xlsx.
"""

import datetime
import re
import os
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────── Paths ───────────────────────
BASE = r"C:\Users\justi\OneDrive\Documents\Claude\Projects\DXY Backtesting"
ASH_PATH = os.path.join(BASE, "Backtesting Results", "Ash Mall DXY - Strategy Results.xlsx")
BRICE_PATH = os.path.join(BASE, "Backtesting Results", "Brice Strebler DXY - Strategy Results.xlsx")
OUTPUT_PATH = os.path.join(BASE, "Backtesting Results", "Phase1_TradeAnalysis.xlsx")

# ─────────────────────── Indicator parameters ───────────────────────
ENTRY_START = 450    # 07:30
ENTRY_END = 1170     # 19:30
MONDAY_START = 390   # 06:30
REV_END = 720        # 12:00
ATTR_MIN_PTS = 1500
ATTR_MAX_PTS = 5000
REV_MIN_SL = 3000
REV_MAX_DIST = 500   # (unused check but defined for reference)

# ─────────────────────── Helpers ───────────────────────

def parse_entry_time(val):
    """Return (hour, minute) from various time formats, or (None, None)."""
    if val is None:
        return None, None
    if isinstance(val, datetime.time):
        return val.hour, val.minute
    if isinstance(val, datetime.datetime):
        return val.hour, val.minute
    # String like "10.30:00 AM", "15:45:00 PM", "08:15:00", "7:00"
    s = str(val).strip()
    # Normalise dot-colon: "10.30:00 AM" -> "10:30:00 AM"
    s = s.replace('.', ':')
    # Try various patterns
    for fmt in ("%I:%M:%S %p", "%I:%M %p", "%H:%M:%S", "%H:%M"):
        try:
            t = datetime.datetime.strptime(s, fmt)
            return t.hour, t.minute
        except ValueError:
            pass
    # Last resort: grab first two numbers
    nums = re.findall(r'\d+', s)
    if len(nums) >= 2:
        h, m = int(nums[0]), int(nums[1])
        # Handle PM indicator
        if 'PM' in s.upper() and h < 12:
            h += 12
        if 0 <= h <= 23 and 0 <= m <= 59:
            return h, m
    return None, None


def time_to_minutes(h, m):
    if h is None:
        return None
    return h * 60 + m


def minutes_to_hhmm(mins):
    if mins is None:
        return ""
    h = int(mins) // 60
    m = int(mins) % 60
    return f"{h:02d}:{m:02d}"


def parse_direction(order_val):
    if order_val is None:
        return "UNKNOWN"
    s = str(order_val).strip().upper()
    if s in ("BUY", "LONG"):
        return "LONG"
    if s in ("SELL", "SHORT"):
        return "SHORT"
    return s


def parse_trade_type(notes):
    if not notes:
        return "UNKNOWN"
    s = str(notes).lower()
    if "attraction" in s:
        return "ATTRACTION"
    if "reversal" in s:
        return "REVERSAL"
    return "UNKNOWN"


def compute_sl_dist(entry, sl):
    try:
        return round(abs(float(entry) - float(sl)) * 10000)
    except (TypeError, ValueError):
        return None


def compute_tp_dist(entry, tp):
    try:
        return round(abs(float(tp) - float(entry)) * 10000)
    except (TypeError, ValueError):
        return None


def compute_rr(tp_dist, sl_dist):
    try:
        if tp_dist and sl_dist and sl_dist > 0:
            return round(tp_dist / sl_dist, 3)
    except (TypeError, ZeroDivisionError):
        pass
    return None


def is_monday(date_val):
    try:
        if isinstance(date_val, (datetime.datetime, datetime.date)):
            return date_val.weekday() == 0  # Monday = 0
    except Exception:
        pass
    return False


def check_in_session(entry_mins, date_val):
    if entry_mins is None:
        return None
    start = MONDAY_START if is_monday(date_val) else ENTRY_START
    return bool(start <= entry_mins <= ENTRY_END)


def check_in_rev_session(entry_mins):
    if entry_mins is None:
        return None
    return bool(entry_mins <= REV_END)


def check_sl_above_rev_min(sl_dist):
    if sl_dist is None:
        return None
    return bool(sl_dist >= REV_MIN_SL)


def check_attr_dist_in_range(trade_type, tp_dist):
    if trade_type != "ATTRACTION":
        return None  # Not applicable
    if tp_dist is None:
        return None
    return bool(ATTR_MIN_PTS <= tp_dist <= ATTR_MAX_PTS)


# ─────────────────────── Load raw data ───────────────────────

def load_sheet(path, sheet="Forex", data_only=True):
    wb = openpyxl.load_workbook(path, data_only=data_only)
    ws = wb[sheet]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return headers, rows


# ─────────────────────── Process one trader's data ───────────────────────

def process_trader(rows, entry_time_col, has_tp_col=False, tp_col=None):
    """
    Filter to DXY wins, sort by Date Entered, take last 50, compute all fields.
    Returns a list of dicts.
    """
    # Filter DXY
    dxy = [r for r in rows if r.get("TICKER") == "DXY"]
    print(f"  Total DXY rows: {len(dxy)}")

    # Filter wins — check exact values
    win_vals = set()
    for r in dxy:
        wl = r.get("Win/Loss")
        if wl is not None:
            win_vals.add(wl)
    print(f"  Win/Loss unique values: {win_vals}")

    wins = [r for r in dxy if r.get("Win/Loss") == "Win"]
    print(f"  DXY Win rows: {len(wins)}")

    # Sort by Date Entered ascending
    def sort_key(r):
        d = r.get("Date Entered")
        if isinstance(d, (datetime.datetime, datetime.date)):
            return d
        return datetime.datetime(1900, 1, 1)

    wins_sorted = sorted(wins, key=sort_key)

    # Take last 50
    last50 = wins_sorted[-50:]
    print(f"  Date range of last 50: {last50[0].get('Date Entered')} to {last50[-1].get('Date Entered')}")

    results = []
    for r in last50:
        trade = {}

        # Raw fields
        trade["#"] = r.get("#")
        trade["Date Entered"] = r.get("Date Entered")
        trade["Date Exit"] = r.get("Date Exit")
        trade["Entry"] = r.get("Entry")
        trade["Stop Loss"] = r.get("Stop Loss")
        trade["Exit price"] = r.get("Exit price")
        trade["Pip Gain"] = r.get("Pip Gain")
        trade["Trade RR"] = r.get("Trade RR")
        trade["% Return"] = r.get("% Return")
        trade["Win/Loss"] = r.get("Win/Loss")
        trade["Duration"] = r.get("Duration")
        trade["Notes"] = r.get("Notes")

        # A. Direction
        trade["Direction"] = parse_direction(r.get("Order"))

        # B. Entry time
        raw_time = r.get(entry_time_col)
        h, m = parse_entry_time(raw_time)
        trade["Entry Time Raw"] = str(raw_time) if raw_time is not None else ""
        trade["Entry Time HH:MM"] = f"{h:02d}:{m:02d}" if h is not None else ""
        entry_mins = time_to_minutes(h, m)
        trade["Entry Time Mins"] = entry_mins

        # C. SL distance
        sl_dist = compute_sl_dist(r.get("Entry"), r.get("Stop Loss"))
        trade["SL Dist (pts)"] = sl_dist

        # D. Trade type
        trade_type = parse_trade_type(r.get("Notes"))
        trade["Trade Type"] = trade_type

        # E. TP distance
        tp_dist = None
        if has_tp_col and tp_col:
            tp_val = r.get(tp_col)
            if tp_val is not None:
                tp_dist = compute_tp_dist(r.get("Entry"), tp_val)
        # For Ash Mall, use Pip Gain as proxy for TP distance
        if tp_dist is None:
            pip_gain = r.get("Pip Gain")
            if pip_gain is not None:
                try:
                    tp_dist = round(abs(float(pip_gain)))
                except (TypeError, ValueError):
                    pass
        trade["TP Dist (pts)"] = tp_dist

        # F. Estimated RR
        trade["Est RR"] = compute_rr(tp_dist, sl_dist)

        # Indicator checks
        trade["in_session"] = check_in_session(entry_mins, r.get("Date Entered"))
        trade["in_rev_session"] = check_in_rev_session(entry_mins)
        trade["sl_above_rev_min"] = check_sl_above_rev_min(sl_dist)
        trade["attr_dist_in_range"] = check_attr_dist_in_range(trade_type, tp_dist)

        results.append(trade)

    return results


# ─────────────────────── Load and process ───────────────────────

print("=" * 60)
print("Loading Ash Mall data...")
ash_headers, ash_rows = load_sheet(ASH_PATH)
print(f"  Headers: {[h for h in ash_headers if h]}")
ash_trades = process_trader(ash_rows, entry_time_col="SHORT", has_tp_col=False)

print()
print("Loading Brice Strebler data...")
brice_headers, brice_rows = load_sheet(BRICE_PATH)
print(f"  Headers: {[h for h in brice_headers if h]}")
# Check for Take Profit column
brice_tp_col = "Take Profit" if "Take Profit" in brice_headers else None
print(f"  Take Profit column: {brice_tp_col}")
brice_trades = process_trader(
    brice_rows,
    entry_time_col="Entry time/pending order bar",
    has_tp_col=(brice_tp_col is not None),
    tp_col=brice_tp_col
)

# ─────────────────────── Summary statistics ───────────────────────

def compute_summary(trades, trader_name):
    df = pd.DataFrame(trades)

    total = len(df)
    summary = {"Trader": trader_name, "Total Trades (last 50 wins)": total}

    # Trade type counts
    for t in ["ATTRACTION", "REVERSAL", "UNKNOWN"]:
        summary[f"Count {t}"] = int((df["Trade Type"] == t).sum())

    # % within session
    in_sess = df["in_session"].dropna()
    summary["% in_session (excl. null)"] = f"{100*in_sess.mean():.1f}%" if len(in_sess) else "N/A"

    # Reversal subset
    rev = df[df["Trade Type"] == "REVERSAL"]
    summary["Reversal trades count"] = len(rev)

    if len(rev) > 0:
        rev_in_sess = rev["in_rev_session"].dropna()
        summary["% reversal in rev_session (<= 12:00)"] = (
            f"{100*rev_in_sess.mean():.1f}%" if len(rev_in_sess) else "N/A"
        )
        rev_sl = rev["sl_above_rev_min"].dropna()
        summary["% reversal SL >= 3000 pts"] = (
            f"{100*rev_sl.mean():.1f}%" if len(rev_sl) else "N/A"
        )
        # >= 5000
        rev_sl_5k = rev["SL Dist (pts)"].dropna()
        pct_5k = (rev_sl_5k >= 5000).mean() if len(rev_sl_5k) else None
        summary["% reversal SL >= 5000 pts"] = f"{100*pct_5k:.1f}%" if pct_5k is not None else "N/A"
    else:
        summary["% reversal in rev_session (<= 12:00)"] = "N/A"
        summary["% reversal SL >= 3000 pts"] = "N/A"
        summary["% reversal SL >= 5000 pts"] = "N/A"

    # SL distance distribution
    sl_vals = df["SL Dist (pts)"].dropna().astype(float)
    if len(sl_vals) > 0:
        summary["SL Dist - Min"] = int(sl_vals.min())
        summary["SL Dist - 25th pct"] = int(np.percentile(sl_vals, 25))
        summary["SL Dist - Median"] = int(np.percentile(sl_vals, 50))
        summary["SL Dist - 75th pct"] = int(np.percentile(sl_vals, 75))
        summary["SL Dist - Max"] = int(sl_vals.max())
    else:
        for k in ["Min", "25th pct", "Median", "75th pct", "Max"]:
            summary[f"SL Dist - {k}"] = "N/A"

    # Entry time distribution (in minutes, displayed as HH:MM)
    et_vals = df["Entry Time Mins"].dropna().astype(float)
    if len(et_vals) > 0:
        summary["Entry Time - Min"] = minutes_to_hhmm(et_vals.min())
        summary["Entry Time - 25th pct"] = minutes_to_hhmm(np.percentile(et_vals, 25))
        summary["Entry Time - Median"] = minutes_to_hhmm(np.percentile(et_vals, 50))
        summary["Entry Time - 75th pct"] = minutes_to_hhmm(np.percentile(et_vals, 75))
        summary["Entry Time - Max"] = minutes_to_hhmm(et_vals.max())
    else:
        for k in ["Min", "25th pct", "Median", "75th pct", "Max"]:
            summary[f"Entry Time - {k}"] = "N/A"

    # Attraction TP distribution
    attr = df[df["Trade Type"] == "ATTRACTION"]
    tp_vals = attr["TP Dist (pts)"].dropna().astype(float)
    summary["Attraction trades count"] = len(attr)
    if len(tp_vals) > 0:
        summary["Attr TP Dist - Min"] = int(tp_vals.min())
        summary["Attr TP Dist - 25th pct"] = int(np.percentile(tp_vals, 25))
        summary["Attr TP Dist - Median"] = int(np.percentile(tp_vals, 50))
        summary["Attr TP Dist - 75th pct"] = int(np.percentile(tp_vals, 75))
        summary["Attr TP Dist - Max"] = int(tp_vals.max())
    else:
        for k in ["Min", "25th pct", "Median", "75th pct", "Max"]:
            summary[f"Attr TP Dist - {k}"] = "N/A"

    return summary


print()
print("Computing summary statistics...")
ash_summary = compute_summary(ash_trades, "Ash Mall")
brice_summary = compute_summary(brice_trades, "Brice Strebler")

# Combined summary
def compute_combined(ash_trades, brice_trades):
    all_trades = ash_trades + brice_trades
    return compute_summary(all_trades, "COMBINED")

combined_summary = compute_combined(ash_trades, brice_trades)

# ─────────────────────── Excel output ───────────────────────

HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
TRUE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FALSE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NONE_FILL = PatternFill(start_color="FFFFC0", end_color="FFFFC0", fill_type="solid")
BOLD = Font(bold=True)

# Column order for trade sheets
TRADE_COLS = [
    "#", "Date Entered", "Date Exit", "Direction", "Entry Time HH:MM",
    "Entry Time Mins", "Entry", "Stop Loss", "Exit price",
    "SL Dist (pts)", "Pip Gain", "TP Dist (pts)", "Est RR",
    "Trade RR", "% Return", "Win/Loss", "Duration", "Trade Type", "Notes",
    "in_session", "in_rev_session", "sl_above_rev_min", "attr_dist_in_range"
]

INDICATOR_COLS = {"in_session", "in_rev_session", "sl_above_rev_min", "attr_dist_in_range"}


def write_trade_sheet(wb, sheet_name, trades):
    ws = wb.create_sheet(title=sheet_name)

    # Write header
    for col_idx, col_name in enumerate(TRADE_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Write data
    for row_idx, trade in enumerate(trades, start=2):
        for col_idx, col_name in enumerate(TRADE_COLS, start=1):
            val = trade.get(col_name)

            # Format date
            if isinstance(val, datetime.datetime):
                val = val.strftime("%Y-%m-%d")

            cell = ws.cell(row=row_idx, column=col_idx, value=val)

            # Colour indicator columns
            if col_name in INDICATOR_COLS:
                if val is True:
                    cell.fill = TRUE_FILL
                elif val is False:
                    cell.fill = FALSE_FILL
                elif val is None:
                    cell.fill = NONE_FILL

    # Auto-fit column widths (approximate)
    for col_idx, col_name in enumerate(TRADE_COLS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(col_name)), 8)
        for row_idx in range(2, len(trades) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, min(len(str(cell_val)), 40))
        ws.column_dimensions[col_letter].width = max_len + 2

    # Freeze top row
    ws.freeze_panes = "A2"

    print(f"  Sheet '{sheet_name}' written with {len(trades)} rows.")
    return ws


def write_summary_sheet(wb, ash_summary, brice_summary, combined_summary):
    ws = wb.create_sheet(title="Summary")

    summaries = [ash_summary, brice_summary, combined_summary]

    # Get all keys in order
    keys = list(ash_summary.keys())

    # Write header row
    ws.cell(row=1, column=1, value="Metric").font = BOLD
    ws.cell(row=1, column=1).fill = HEADER_FILL
    for col_idx, s in enumerate(summaries, start=2):
        cell = ws.cell(row=1, column=col_idx, value=s["Trader"])
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Write each metric row
    for row_idx, key in enumerate(keys, start=2):
        if key == "Trader":
            continue
        ws.cell(row=row_idx, column=1, value=key)
        for col_idx, s in enumerate(summaries, start=2):
            ws.cell(row=row_idx, column=col_idx, value=s.get(key, ""))

    # Auto-fit columns
    for col_idx in range(1, len(summaries) + 2):
        col_letter = get_column_letter(col_idx)
        max_len = 10
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 50))
        ws.column_dimensions[col_letter].width = max_len + 2

    print(f"  Sheet 'Summary' written.")
    return ws


print()
print("Writing Excel output...")
wb = Workbook()
# Remove default sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

write_trade_sheet(wb, "Ash Mall Last 50", ash_trades)
write_trade_sheet(wb, "Brice Last 50", brice_trades)
write_summary_sheet(wb, ash_summary, brice_summary, combined_summary)

wb.save(OUTPUT_PATH)
print(f"\nSaved: {OUTPUT_PATH}")

# ─────────────────────── Text report ───────────────────────

def print_summary(s):
    print(f"\n{'='*50}")
    print(f"  TRADER: {s['Trader']}")
    print(f"{'='*50}")
    for k, v in s.items():
        if k == "Trader":
            continue
        print(f"  {k:<45} {v}")


print("\n\n" + "=" * 60)
print("PHASE 1 TRADE ANALYSIS — DETAILED TEXT REPORT")
print("=" * 60)

print_summary(ash_summary)
print_summary(brice_summary)
print_summary(combined_summary)

# Detailed per-trade entry time stats
print("\n\n--- ASH MALL LAST 50: Entry Times ---")
for t in ash_trades:
    print(f"  #{t['#']:>5}  {str(t.get('Date Entered',''))[:10]}  "
          f"{t.get('Entry Time HH:MM','?'):>5}  {t.get('Direction','?'):>5}  "
          f"SL:{str(t.get('SL Dist (pts)','?')):>5}  "
          f"Type:{t.get('Trade Type','?'):>12}  "
          f"in_sess:{str(t.get('in_session','?')):>5}  "
          f"in_rev:{str(t.get('in_rev_session','?')):>5}  "
          f"sl_rev_min:{str(t.get('sl_above_rev_min','?')):>5}")

print("\n--- BRICE LAST 50: Entry Times ---")
for t in brice_trades:
    print(f"  #{t['#']:>5}  {str(t.get('Date Entered',''))[:10]}  "
          f"{t.get('Entry Time HH:MM','?'):>5}  {t.get('Direction','?'):>5}  "
          f"SL:{str(t.get('SL Dist (pts)','?')):>5}  "
          f"Type:{t.get('Trade Type','?'):>12}  "
          f"in_sess:{str(t.get('in_session','?')):>5}  "
          f"in_rev:{str(t.get('in_rev_session','?')):>5}  "
          f"sl_rev_min:{str(t.get('sl_above_rev_min','?')):>5}")

# ─────────────────────── recalc.py ───────────────────────
RECALC_PATH = os.path.join(BASE, "scripts", "recalc.py")
if os.path.exists(RECALC_PATH):
    print(f"\nRunning recalc.py...")
    import subprocess
    result = subprocess.run(
        ["python", RECALC_PATH, OUTPUT_PATH],
        capture_output=True, text=True
    )
    print(result.stdout)
    if result.stderr:
        print("STDERR:", result.stderr)
else:
    print(f"\nNo recalc.py found at {RECALC_PATH} — skipping.")
