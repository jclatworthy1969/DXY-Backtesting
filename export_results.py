import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

with open(r'C:\Users\justi\OneDrive\Documents\Claude\Projects\DXY Backtesting\zone_match_results.json') as f:
    results = json.load(f)

OUT  = r'C:\Users\justi\OneDrive\Desktop\Claude is a genius.xlsx'
RISK = 1000

# ── Feb-Mar 2026 forward test data (from Pine Script indicator simulation) ────
FWD = {
    'period':        'Feb 1 – Mar 31, 2026',
    'total_trades':  17,
    'wins':          3,
    'losses':        14,
    'win_rate':      17.6,
    'avg_dur_min':   221,       # rounded
    'win_pips':      606,       # DXY pips (÷10 from indicator units)
    'loss_pips':     2478,
    'net_pips':      -1872,
    'total_r':       -11,
    'pnl_usd':       -11000,
    'pct_return':    -11.0,
    'trade_types': {
        'REV SHORT': 9,
        'REV LONG':  5,
        'ATTR LONG': 2,
        'ATTR SHORT':1,
    }
}

# ── Colours ───────────────────────────────────────────────────────────────────
HDR_BG    = '1F3864'
HDR_FG    = 'FFFFFF'
SUB_BG    = '2E75B6'
ATTR_BG   = 'E2EFDA'
REV_BG    = 'DEEAF1'
MANUAL_BG = 'FFF2CC'
OTHER_BG  = 'FCE4D6'
ALT_BG    = 'F5F5F5'
WIN_BG    = 'C6EFCE'; WIN_FG = '276221'
LOSS_BG   = 'FFC7CE'; LOSS_FG= '9C0006'
SUMM_BG   = 'D9E1F2'
TITLE_BG  = '1F3864'
FWD_BG    = '2D5016'   # dark green for forward test header
FWD_SUB   = '375623'

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def apply_hdr(cell, text, bg=HDR_BG, fg=HDR_FG, size=11, bold=True, wrap=False):
    cell.value = text
    cell.font = Font(name='Arial', bold=bold, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    cell.border = thin_border()

def apply_cell(cell, value, fmt=None, bold=False, align='center', bg=None, fg='000000', wrap=False):
    cell.value = value
    cell.font = Font(name='Arial', bold=bold, color=fg, size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border = thin_border()
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    if fmt:
        cell.number_format = fmt

wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'Summary'
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 36
ws.column_dimensions['B'].width = 24
ws.column_dimensions['C'].width = 24
ws.column_dimensions['D'].width = 24

# Title
ws.merge_cells('A1:D1')
c = ws['A1']
c.value = 'DXY Zone Strategy — Backtest & Forward Test Results'
c.font = Font(name='Arial', bold=True, size=16, color=HDR_FG)
c.fill = PatternFill('solid', fgColor=TITLE_BG)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 36

ws.merge_cells('A2:D2')
c = ws['A2']
c.value = '$1,000 risk per trade  |  1:1 Risk-Reward  |  All times UTC'
c.font = Font(name='Arial', size=11, color='666666', italic=True)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 22

# ── Section 1: Historical Classification ─────────────────────────────────────
ws.row_dimensions[3].height = 8
ws.merge_cells('A4:D4')
apply_hdr(ws['A4'], 'SECTION 1 — HISTORICAL TEST: May–Dec 2024  (Last 50 Winning Trades)', bg=SUB_BG)
ws.row_dimensions[4].height = 22

for col, h in enumerate(['Category', 'Count', '% of 50 Trades', 'Notes'], 1):
    apply_hdr(ws.cell(5, col), h, bg='2F5496', size=10)
ws.row_dimensions[5].height = 20

match_counts = Counter(r['match'] for r in results)
categories = [
    ('ATTRACTION',           ATTR_BG,   'Body-clean zone, signal, 150–500 DXY pts to TP'),
    ('REVERSAL',             REV_BG,    'Broken zone as S/R, any entry signal, in session'),
    ('MANUAL_CHECK',         MANUAL_BG, 'Missing zone data — manual chart review required'),
    ('TOO_CLOSE_TO_ZONE',    OTHER_BG,  'Entry < 150 pts from zone far-side TP'),
    ('BODY_IN_ZONE_NO_ATTR', OTHER_BG,  'Body inside zone — attraction condition failed'),
]
row = 6
for cat, bg, note in categories:
    cnt = match_counts.get(cat, 0)
    apply_cell(ws.cell(row, 1), cat, bold=True, align='left', bg=bg)
    apply_cell(ws.cell(row, 2), cnt, fmt='0', bg=bg)
    apply_cell(ws.cell(row, 3), cnt/50, fmt='0%', bg=bg)
    apply_cell(ws.cell(row, 4), note, align='left', bg=bg, wrap=True)
    ws.row_dimensions[row].height = 20
    row += 1

matched   = [r for r in results if r['match'] in ('ATTRACTION', 'REVERSAL')]
attr_t    = [r for r in matched if r['match'] == 'ATTRACTION']
rev_t     = [r for r in matched if r['match'] == 'REVERSAL']

def R(t):   return t['pip_gain'] / t['sl_gap_pts'] if t['sl_gap_pts'] else 0
def pnl(t): return R(t) * RISK

total_pnl  = sum(pnl(t) for t in matched)
avg_r      = sum(R(t) for t in matched) / len(matched)
avg_pips   = sum(t['pip_gain'] for t in matched) / len(matched)
attr_avg_r = sum(R(t) for t in attr_t) / len(attr_t) if attr_t else 0
rev_avg_r  = sum(R(t) for t in rev_t)  / len(rev_t)  if rev_t  else 0

ws.row_dimensions[row].height = 8; row += 1
for col, h in enumerate(['Metric', 'Historical (May–Dec 2024)', '', ''], 1):
    apply_hdr(ws.cell(row, col), h, bg='2F5496', size=10)
ws.row_dimensions[row].height = 20; row += 1

hist_rows = [
    ('Matched trades (of 50)',       f'{len(matched)} trades ({len(attr_t)} attr, {len(rev_t)} rev)', '', ''),
    ('Win rate (pre-screened wins)',  '100%',                          '',  'Dataset = confirmed winners only'),
    ('Total P&L  @ $1,000/trade',    f'${total_pnl:,.0f}',            '',  ''),
    ('Avg R per trade',              f'{avg_r:.2f}R',                  '',  ''),
    ('Avg pip gain',                 f'{avg_pips:.0f} pts',            '',  f'≈ {avg_pips/10:.0f} DXY pips'),
]
for label, val, sub, note in hist_rows:
    apply_cell(ws.cell(row, 1), label, align='left', bg=SUMM_BG, bold=True)
    apply_cell(ws.cell(row, 2), val,   bg=SUMM_BG)
    apply_cell(ws.cell(row, 3), sub,   bg=SUMM_BG)
    apply_cell(ws.cell(row, 4), note,  align='left', bg=SUMM_BG, wrap=True)
    ws.row_dimensions[row].height = 20; row += 1

# ── Section 2: Feb-Mar 2026 Forward Test ─────────────────────────────────────
ws.row_dimensions[row].height = 10; row += 1
ws.merge_cells(f'A{row}:D{row}')
apply_hdr(ws[f'A{row}'], 'SECTION 2 — FORWARD TEST: Feb–Mar 2026  (All Raw Signals)', bg='375623')
ws.row_dimensions[row].height = 22; row += 1

for col, h in enumerate(['Metric', 'Forward Test (Feb–Mar 2026)', 'vs Historical', 'Notes'], 1):
    apply_hdr(ws.cell(row, col), h, bg='4E7A1E', size=10)
ws.row_dimensions[row].height = 20; row += 1

fwd_rows = [
    ('Period',             FWD['period'],                                  '',          '59 calendar days'),
    ('Total Trades',       str(FWD['total_trades']),                       '',          '~2 trades/week'),
    ('Wins / Losses',      f"{FWD['wins']} wins / {FWD['losses']} losses", '',          ''),
    ('Win Rate',           f"{FWD['win_rate']}%",                          '↓ vs 100%', 'All raw signals, not pre-screened'),
    ('Avg Trade Duration', f"{FWD['avg_dur_min']} min  (3h 41min)",        '',          ''),
    ('Win Pips',           f"+{FWD['win_pips']:,} DXY pips",               '',          f"{FWD['wins']} winning trades"),
    ('Loss Pips',          f"-{FWD['loss_pips']:,} DXY pips",              '',          f"{FWD['losses']} losing trades"),
    ('Net Pips',           f"{FWD['net_pips']:,} DXY pips",                '',          ''),
    ('Total R',            f"{FWD['total_r']}R",                           '',          ''),
    ('P&L @ $1,000/trade', f"${FWD['pnl_usd']:,}",                        '',          ''),
    ('% Return ($100k)',   f"{FWD['pct_return']}%",                        '',          '1% risk per trade on $100k capital'),
]
for label, val, vs, note in fwd_rows:
    is_pnl  = 'P&L' in label or '%' in label or 'Net' in label or 'Win Rate' in label
    val_bg  = (LOSS_BG if (isinstance(val, str) and val.startswith('-')) or val == f"{FWD['pct_return']}%" or val == f"{FWD['total_r']}R" or val == f"${FWD['pnl_usd']:,}"
               else (WIN_BG if isinstance(val, str) and val.startswith('+') else ALT_BG))
    val_fg  = (LOSS_FG if val_bg == LOSS_BG else (WIN_FG if val_bg == WIN_BG else '000000'))
    apply_cell(ws.cell(row, 1), label, align='left', bg=ALT_BG, bold=True)
    apply_cell(ws.cell(row, 2), val,   bg=val_bg, fg=val_fg, bold=is_pnl)
    apply_cell(ws.cell(row, 3), vs,    bg=ALT_BG, fg='7F7F7F')
    apply_cell(ws.cell(row, 4), note,  align='left', bg=ALT_BG, wrap=True)
    ws.row_dimensions[row].height = 20; row += 1

ws.row_dimensions[row].height = 8; row += 1
ws.merge_cells(f'A{row}:D{row}')
c = ws[f'A{row}']
c.value = ('⚠  Forward test uses process_orders_on_close semantics (fills at bar close). '
           'DXY was in a strong downtrend (108 → 98) during Feb–Mar 2026, which challenged '
           'the zone strategy. Reversal trades made up 14 of 17 signals. '
           'Historical test uses confirmed winners only — not directly comparable to raw forward test.')
c.font = Font(name='Arial', size=9, color='7F7F7F', italic=True)
c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws.row_dimensions[row].height = 40

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — TRADE LOG (historical matched trades)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Historical Trade Log')
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = 'A3'

col_widths = [6, 12, 8, 12, 10, 10, 10, 10, 10, 10, 10, 8, 12, 28]
col_headers = [
    '#', 'Date', 'Dir', 'Setup', 'Entry', 'Stop Loss',
    'Exit', 'SL Gap\n(pts)', 'Pip Gain\n(pts)', 'R Multiple',
    '$P&L\n@$1k', 'Session', 'Dist Label', 'Zone Details'
]
for i, w in enumerate(col_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells('A1:N1')
c = ws2['A1']
c.value = 'DXY Zone Strategy — Historical Matched Trade Log (May–Dec 2024)'
c.font = Font(name='Arial', bold=True, size=14, color=HDR_FG)
c.fill = PatternFill('solid', fgColor=TITLE_BG)
c.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 32

for col, h in enumerate(col_headers, 1):
    apply_hdr(ws2.cell(2, col), h, wrap=True)
ws2.row_dimensions[2].height = 32

for i, t in enumerate(sorted(matched, key=lambda x: x['date']), 1):
    row = i + 2
    r_val   = R(t)
    pnl_val = r_val * RISK
    bg      = ATTR_BG if t['match'] == 'ATTRACTION' else REV_BG
    zone_str = (f"{t['zone_bottom']:.3f}–{t['zone_top']:.3f}  "
                f"({'BULL' if t['japan_bull'] else 'BEAR'})  "
                f"P={int(t['pristine'])}  BC={int(t['body_clean'])}")

    apply_cell(ws2.cell(row, 1),  i,              fmt='0',         bg=bg)
    apply_cell(ws2.cell(row, 2),  t['date'],       bg=bg)
    dir_c = ws2.cell(row, 3)
    dir_c.value = t['direction']
    dir_c.font  = Font(name='Arial', bold=True, size=10,
                       color=WIN_FG if t['direction']=='LONG' else LOSS_FG)
    dir_c.fill  = PatternFill('solid', fgColor=WIN_BG if t['direction']=='LONG' else LOSS_BG)
    dir_c.alignment = Alignment(horizontal='center', vertical='center')
    dir_c.border = thin_border()
    apply_cell(ws2.cell(row, 4),  t['match'],      bold=True, bg=bg)
    apply_cell(ws2.cell(row, 5),  t['entry_px'],   fmt='#,##0.000', bg=bg)
    apply_cell(ws2.cell(row, 6),  t['sl'],         fmt='#,##0.000', bg=bg)
    apply_cell(ws2.cell(row, 7),  t['exit_px'],    fmt='#,##0.000', bg=bg)
    apply_cell(ws2.cell(row, 8),  t['sl_gap_pts'], fmt='#,##0',     bg=bg)
    apply_cell(ws2.cell(row, 9),  t['pip_gain'],   fmt='#,##0',     bg=bg)
    apply_cell(ws2.cell(row, 10), r_val,           fmt='0.00"R"',   bg=bg)
    pnl_c = ws2.cell(row, 11)
    pnl_c.value = pnl_val
    pnl_c.font  = Font(name='Arial', bold=True, size=10, color=WIN_FG)
    pnl_c.fill  = PatternFill('solid', fgColor=WIN_BG)
    pnl_c.number_format = '"$"#,##0'
    pnl_c.alignment = Alignment(horizontal='center', vertical='center')
    pnl_c.border = thin_border()
    apply_cell(ws2.cell(row, 12), 'Yes' if t['in_sess'] else 'No', bg=bg)
    apply_cell(ws2.cell(row, 13), t.get('dist_label',''), align='left', bg=bg)
    apply_cell(ws2.cell(row, 14), zone_str, align='left', bg=bg, wrap=True)
    ws2.row_dimensions[row].height = 18

tot_row = len(matched) + 3
ws2.merge_cells(f'A{tot_row}:I{tot_row}')
apply_hdr(ws2[f'A{tot_row}'], f'TOTAL  ({len(matched)} matched trades)', bg=HDR_BG)
apply_hdr(ws2.cell(tot_row, 10), f'{avg_r:.2f}R avg', bg=HDR_BG)
apply_hdr(ws2.cell(tot_row, 11), f'${total_pnl:,.0f}', bg='276221')
for col in range(12, 15):
    apply_hdr(ws2.cell(tot_row, col), '', bg=HDR_BG)
ws2.row_dimensions[tot_row].height = 22

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — FEB-MAR 2026 FORWARD TEST DETAILS
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Feb-Mar 2026 Forward Test')
ws3.sheet_view.showGridLines = False

col_w3 = [6, 20, 22, 16, 16, 16]
for i, w in enumerate(col_w3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells('A1:F1')
c = ws3['A1']
c.value = 'DXY Zone Strategy — Forward Test: Feb 1 – Mar 31, 2026'
c.font = Font(name='Arial', bold=True, size=14, color=HDR_FG)
c.fill = PatternFill('solid', fgColor=FWD_BG)
c.alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 32

ws3.merge_cells('A2:F2')
c = ws3['A2']
c.value = '17 trades  |  $1,000 risk per trade  |  1:1 RR  |  process_orders_on_close semantics  |  DXY 15m TVC:DXY'
c.font = Font(name='Arial', size=10, color='AAAAAA', italic=True)
c.fill = PatternFill('solid', fgColor='1A1A1A')
c.alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[2].height = 20

# KPI summary boxes
ws3.row_dimensions[3].height = 8
for col, h in enumerate(['KPI', 'Value', 'Trade Type', 'Count', 'Win Rate', 'Context'], 1):
    apply_hdr(ws3.cell(4, col), h, bg='4E7A1E', size=10)
ws3.row_dimensions[4].height = 20

kpi_rows = [
    ('Total Trades',         '17',                        'REV SHORT',  '9',  '',      '53% of all trades'),
    ('Wins / Losses',        '3 / 14',                    'REV LONG',   '5',  '',      '29% of all trades'),
    ('Win Rate',             '17.6%',                     'ATTR LONG',  '2',  '',      '12% of all trades'),
    ('Avg Duration',         '221 min  (3h 41min)',        'ATTR SHORT', '1',  '',      '6% of all trades'),
    ('Win Pips (gross)',     '+606 DXY pips',              '',           '',   '',      'Avg +202 pips/win'),
    ('Loss Pips (gross)',    '-2,478 DXY pips',            '',           '',   '',      'Avg -177 pips/loss'),
    ('Net Pips',            '-1,872 DXY pips',             '',           '',   '',      ''),
    ('Total R',             '-11R',                        '',           '',   '',      '3 wins × +1R, 14 losses × -1R'),
    ('P&L @ $1,000/trade',  '-$11,000',                    '',           '',   '',      ''),
    ('% Return on $100k',   '-11.0%',                      '',           '',   '',      '1% risk per trade'),
]
for r_idx, (kpi, val, ttype, cnt, wr, ctx) in enumerate(kpi_rows, 5):
    is_neg = val.startswith('-') if isinstance(val, str) else False
    is_pos = val.startswith('+') if isinstance(val, str) else False
    val_bg = LOSS_BG if is_neg else (WIN_BG if is_pos else ALT_BG)
    val_fg = LOSS_FG if is_neg else (WIN_FG if is_pos else '000000')
    apply_cell(ws3.cell(r_idx, 1), kpi,   align='left',  bg=ALT_BG, bold=True)
    apply_cell(ws3.cell(r_idx, 2), val,   align='center', bg=val_bg, fg=val_fg, bold=True)
    apply_cell(ws3.cell(r_idx, 3), ttype, align='left',  bg=SUMM_BG)
    apply_cell(ws3.cell(r_idx, 4), cnt,   align='center', bg=SUMM_BG)
    apply_cell(ws3.cell(r_idx, 5), wr,    align='center', bg=SUMM_BG)
    apply_cell(ws3.cell(r_idx, 6), ctx,   align='left',  bg=ALT_BG, wrap=True)
    ws3.row_dimensions[r_idx].height = 22

# Entry price list from pine labels
ws3.row_dimensions[15].height = 8
for col, h in enumerate(['#', 'Trade Type', 'Direction', 'Entry Price (DXY)', 'Result', 'Notes'], 1):
    apply_hdr(ws3.cell(16, col), h, bg='375623', size=10)
ws3.row_dimensions[16].height = 20

pine_labels = [
    (1,  'REV SHORT',  'SHORT', 97.290),
    (2,  'REV SHORT',  'SHORT', 97.030),
    (3,  'REV SHORT',  'SHORT', 97.550),
    (4,  'REV LONG',   'LONG',  98.150),
    (5,  'REV LONG',   'LONG',  99.120),
    (6,  'REV LONG',   'LONG',  98.560),
    (7,  'REV SHORT',  'SHORT', 98.980),
    (8,  'REV SHORT',  'SHORT', 100.160),
    (9,  'ATTR LONG',  'LONG',  99.930),
    (10, 'REV SHORT',  'SHORT', 99.580),
    (11, 'ATTR LONG',  'LONG',  99.670),
    (12, 'ATTR SHORT', 'SHORT', 99.970),
    (13, 'REV SHORT',  'SHORT', 99.290),
    (14, 'REV LONG',   'LONG',  99.270),
    (15, 'REV SHORT',  'SHORT', 99.770),
    (16, 'REV SHORT',  'SHORT', 100.060),
    (17, 'REV LONG',   'LONG',  100.370),
]
for num, ttype, dirn, entry in pine_labels:
    row_n = num + 16
    bg = ATTR_BG if 'ATTR' in ttype else REV_BG
    apply_cell(ws3.cell(row_n, 1), num,   fmt='0',         bg=bg)
    apply_cell(ws3.cell(row_n, 2), ttype, bold=True,       bg=bg)
    dir_c = ws3.cell(row_n, 3)
    dir_c.value = dirn
    dir_c.font  = Font(name='Arial', bold=True, size=10,
                       color=WIN_FG if dirn=='LONG' else LOSS_FG)
    dir_c.fill  = PatternFill('solid', fgColor=WIN_BG if dirn=='LONG' else LOSS_BG)
    dir_c.alignment = Alignment(horizontal='center', vertical='center')
    dir_c.border = thin_border()
    apply_cell(ws3.cell(row_n, 4), entry, fmt='#,##0.000', bg=bg)
    apply_cell(ws3.cell(row_n, 5), 'W or L',              bg=ALT_BG, fg='888888')
    apply_cell(ws3.cell(row_n, 6), 'From pine labels — exact outcome per-trade requires chart review',
               align='left', bg=ALT_BG, fg='888888', wrap=True)
    ws3.row_dimensions[row_n].height = 18

ws3.row_dimensions[34].height = 8
ws3.merge_cells('A35:F35')
c = ws3['A35']
c.value = ('Note: DXY was in a strong downtrend (≈108 → 98) during Feb–Mar 2026. '
           'Reversal trades dominated (14/17). Sim uses bar-close fills (conservative). '
           'Attraction trades: 3, Reversal trades: 14.')
c.font = Font(name='Arial', size=9, color='7F7F7F', italic=True)
c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws3.row_dimensions[35].height = 36

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — ALL 50 HISTORICAL TRADES
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('All 50 Historical Trades')
ws4.sheet_view.showGridLines = False
ws4.freeze_panes = 'A3'

col_w4 = [6, 12, 8, 20, 10, 10, 10, 10, 32]
col_h4 = ['#', 'Date', 'Dir', 'Classification', 'Entry', 'SL', 'Exit', 'Pip Gain', 'Zone / Notes']
for i, w in enumerate(col_w4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.merge_cells('A1:I1')
c = ws4['A1']
c.value = 'DXY Zone Strategy — All 50 Historical Backtested Trades (May–Dec 2024)'
c.font = Font(name='Arial', bold=True, size=14, color=HDR_FG)
c.fill = PatternFill('solid', fgColor=TITLE_BG)
c.alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 32

for col, h in enumerate(col_h4, 1):
    apply_hdr(ws4.cell(2, col), h)
ws4.row_dimensions[2].height = 20

cat_bg = {
    'ATTRACTION':           ATTR_BG,
    'REVERSAL':             REV_BG,
    'MANUAL_CHECK':         MANUAL_BG,
    'TOO_CLOSE_TO_ZONE':    OTHER_BG,
    'BODY_IN_ZONE_NO_ATTR': OTHER_BG,
}
for i, t in enumerate(sorted(results, key=lambda x: x['date']), 1):
    row = i + 2
    bg  = cat_bg.get(t['match'], ALT_BG)
    apply_cell(ws4.cell(row, 1), i, fmt='0', bg=bg)
    apply_cell(ws4.cell(row, 2), t['date'], bg=bg)
    dir_c = ws4.cell(row, 3)
    dir_c.value = t['direction']
    dir_c.font  = Font(name='Arial', bold=True, size=10,
                       color=WIN_FG if t['direction']=='LONG' else LOSS_FG)
    dir_c.fill  = PatternFill('solid', fgColor=WIN_BG if t['direction']=='LONG' else LOSS_BG)
    dir_c.alignment = Alignment(horizontal='center', vertical='center')
    dir_c.border = thin_border()
    apply_cell(ws4.cell(row, 4), t['match'], bold=True, bg=bg)
    apply_cell(ws4.cell(row, 5), t['entry_px'],          fmt='#,##0.000', bg=bg)
    apply_cell(ws4.cell(row, 6), t.get('sl', ''),        fmt='#,##0.000', bg=bg)
    apply_cell(ws4.cell(row, 7), t.get('exit_px', ''),   fmt='#,##0.000', bg=bg)
    apply_cell(ws4.cell(row, 8), t.get('pip_gain', ''),  fmt='#,##0',     bg=bg)
    if t.get('zone_found'):
        note = (f"Zone: {t['zone_bottom']:.3f}–{t['zone_top']:.3f}  "
                f"({'BULL' if t['japan_bull'] else 'BEAR'})  "
                f"P={int(t['pristine'])} BC={int(t['body_clean'])}  |  {t.get('dist_label','')}")
    else:
        note = 'No zone data — manual chart review required'
    apply_cell(ws4.cell(row, 9), note, align='left', bg=bg, wrap=True)
    ws4.row_dimensions[row].height = 18

wb.save(OUT)
print(f'Saved: {OUT}')
